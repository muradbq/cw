from openpyxl import load_workbook

# Загрузите файл
file_path = "static/LOGO_RECIPE_TEMPLATE.xlsx"
workbook = load_workbook(filename=file_path)
sheet = workbook.active

headers = [sheet[f"{col}1"].value for col in "ABCDEFGHIJK"]

check = ['Satır Tipi', 'Malzeme Kodu', 'Malzeme Açıklaması', 'Açıklama 2', 'Miktar',
         'Birim', 'Fire Faktörü', 'Net Ağırlık', 'Diametr', 'Operasyon Kodu', 'Operasyon Süresi']

print(headers)

if headers != check:
    print("ERROR")

# Считываем значения из колонок A-K начиная со второй строки
data = []
for row in sheet.iter_rows(min_row=2, max_col=11, values_only=True):
    data.append(list(row))

print('data: ', data)
print('row: ', row)

# Печатаем извлеченные данные
# for row in data:
    # print(row)
